"""자동 CH1.xlsx 파서 — 두 시트 → 두 테이블.

전략: ETL은 xlsx → DB 단순 복사. 해석은 agent.

xlsx 구조:
- 'Main Process' 시트: 38 컬럼 → sputter_runs_auto_main
- 'Plasma Cleaning' 시트: 16 컬럼 → sputter_runs_auto_plasma
- 두 시트 공통: R1=카테고리(skip), R2=헤더, R3=단위, R4부터 데이터
- 첫 컬럼(날짜) 비면 데이터 끝 (자동 기록이라 깨끗)
"""

import json
import logging
from datetime import datetime, date

from openpyxl import load_workbook
from sqlalchemy import text

from shared.db import session_scope_writer
from etl_worker.jobs.scan_files import SourceFileRecord
from etl_worker.jobs.parsers._incremental import row_number_watermark

log = logging.getLogger("etl.parsers.sputter_auto")

# 'Main Process' 시트 (38 컬럼)
# (col_idx, db_col, type)  type: 'timestamp' | 'real' | 'integer' | 'text'
MAIN_COL_MAP = [
    (0,  'process_datetime',   'timestamp'),
    (1,  'operator',           'text'),
    (2,  'process_name',       'text'),
    (3,  'notes',              'text'),
    (4,  'substrate',          'text'),
    (5,  'main_shutter',       'text'),
    (6,  'power_select',       'text'),
    (7,  'g1_target',          'text'),
    (8,  'g2_target',          'text'),
    (9,  'g3_target',          'text'),
    (10, 'deposition_rate',    'real'),
    (11, 'thickness_nm',       'integer'),
    (12, 'chuck',              'text'),
    (13, 'shutter_delay_min',  'real'),
    (14, 'process_time_min',   'real'),
    (15, 'base_pressure_torr', 'real'),
    (16, 'sp_ar_sccm',         'integer'),
    (17, 'avg_ar_sccm',        'real'),
    (18, 'sp_n2_sccm',         'integer'),
    (19, 'avg_n2_sccm',        'real'),
    (20, 'sp_o2_sccm',         'integer'),
    (21, 'avg_o2_sccm',        'real'),
    (22, 'sp_pressure_mtorr',  'integer'),
    (23, 'avg_pressure_mtorr', 'real'),
    (24, 'power_source',       'text'),
    (25, 'sp_power_w',         'integer'),
    (26, 'avg_power_w',        'integer'),
    (27, 'avg_for_p_w',        'real'),
    (28, 'avg_ref_p_w',        'real'),
    (29, 'avg_load',           'text'),
    (30, 'avg_tune',           'text'),
    (31, 'avg_voltage_v',      'real'),
    (32, 'avg_current_a',      'real'),
    (33, 'duty_cycle_pct',     'integer'),
    (34, 'frequency_khz',      'integer'),
    (35, 'off_time_us',        'integer'),
    (36, 'soft_arc_count',     'integer'),
    (37, 'hard_arc_count',     'integer'),
]

# 'Plasma Cleaning' 시트 (16 컬럼)
PLASMA_COL_MAP = [
    (0,  'process_datetime',   'timestamp'),
    (1,  'operator',           'text'),
    (2,  'process_name',       'text'),
    (3,  'notes',              'text'),
    (4,  'substrate',          'text'),
    (5,  'time_min',           'real'),
    (6,  'base_pressure_torr', 'real'),
    (7,  'sp_ar_sccm',         'integer'),
    (8,  'avg_ar_sccm',        'real'),
    (9,  'sp_pressure_mtorr',  'integer'),
    (10, 'avg_pressure_mtorr', 'real'),
    (11, 'sp_power_w',         'integer'),
    (12, 'avg_for_p_w',        'real'),
    (13, 'avg_ref_p_w',        'real'),
    (14, 'avg_load',           'text'),
    (15, 'avg_tune',           'text'),
]


def _to_real(v):
    if v is None or v == '':
        return None
    if isinstance(v, str) and not v.strip():
        return None
    try:
        f = float(v)
        if f != f or f == float('inf') or f == float('-inf'):
            return None
        return f
    except (ValueError, TypeError):
        return None


def _to_int(v):
    if v is None or v == '':
        return None
    if isinstance(v, str) and not v.strip():
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def _to_text(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return str(v)


def _to_timestamp(v):
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime.combine(v, datetime.min.time())
    return None


def _serialize(v):
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return v


def _row_to_full_dict(row, header):
    result = {}
    for i, val in enumerate(row):
        key = header[i] if i < len(header) and header[i] else f'col_{i}'
        if isinstance(key, str):
            key = key.replace('\n', ' ').strip()
        result[key] = val
    return result


def _build_payload(row, col_map, source_file_id, row_number, header):
    payload = {}
    for col_idx, db_col, dtype in col_map:
        v = row[col_idx] if col_idx < len(row) else None
        if dtype == 'real':
            payload[db_col] = _to_real(v)
        elif dtype == 'integer':
            payload[db_col] = _to_int(v)
        elif dtype == 'timestamp':
            payload[db_col] = _to_timestamp(v)
        else:
            payload[db_col] = _to_text(v)

    payload['source_file_id'] = source_file_id
    payload['row_number'] = row_number
    payload['parse_status'] = None

    raw = _row_to_full_dict(row, header)
    payload['raw_json'] = json.dumps(
        {k: _serialize(v) for k, v in raw.items()},
        ensure_ascii=False,
    )
    return payload


def _build_insert_sql(table_name, col_map):
    """col_map → INSERT SQL 동적 생성."""
    db_cols = [m[1] for m in col_map]
    return text(f"""
        INSERT INTO vo2.{table_name} (
            {', '.join(db_cols)},
            source_file_id, row_number, raw_json, parse_status
        ) VALUES (
            {', '.join(':' + c for c in db_cols)},
            :source_file_id, :row_number, CAST(:raw_json AS JSONB), :parse_status
        )
        ON CONFLICT (source_file_id, row_number) DO NOTHING
    """)


_INSERT_MAIN_SQL = _build_insert_sql('sputter_runs_auto_main', MAIN_COL_MAP)
_INSERT_PLASMA_SQL = _build_insert_sql('sputter_runs_auto_plasma', PLASMA_COL_MAP)

_UPDATE_METADATA_SQL = text("""
    UPDATE vo2.source_files
    SET metadata = CAST(:metadata AS JSONB),
        row_count = :row_count,
        parser_status = :parser_status,
        parser_error = :parser_error,
        last_indexed_at = NOW()
    WHERE id = :id
""")


def _process_sheet(ws, col_map, insert_sql, source_file_id, table_name):
    """단일 시트 처리. (inserted, skipped_old) 반환.

    데이터 시작: R4 (R1=카테고리, R2=헤더, R3=단위)
    빈 row 판정: 첫 컬럼(날짜) 비면 데이터 끝

    Args:
        table_name: 'sputter_runs_auto_main' 또는 'sputter_runs_auto_plasma'
                    Step 23 watermark 조회용.
    """
    inserted = 0
    skipped_old = 0

    # Step 23: incremental
    watermark = row_number_watermark(f"vo2.{table_name}")

    # R2 헤더 (raw_json 키용)
    header_rows = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    header = list(header_rows[0]) if header_rows else []
    header = [
        (h.replace('\n', ' ').strip() if isinstance(h, str) else h)
        for h in header
    ]

    with session_scope_writer() as s:
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=4, values_only=True), start=4
        ):
            first_cell = row[0] if row else None
            if first_cell is None or (
                isinstance(first_cell, str) and not first_cell.strip()
            ):
                break

            # Step 23: incremental skip
            if row_idx <= watermark:
                skipped_old += 1
                continue

            payload = _build_payload(row, col_map, source_file_id, row_idx, header)
            s.execute(insert_sql, payload)
            inserted += 1

    log.info(
        f"sputter_auto _process_sheet({table_name}): "
        f"+{inserted} rows, {skipped_old} skipped, watermark={watermark}"
    )
    return inserted, skipped_old


def parse_sputter_auto(record: SourceFileRecord) -> dict:
    """자동 CH1.xlsx 메인 진입. 두 시트 처리.

    Returns:
        {"status": ..., "main_inserted": N, "plasma_inserted": M, ...}
    """
    if record.is_race_unsafe:
        log.info(f"skip {record.file_name} (race_unsafe)")
        return {
            "status": "skipped", "reason": "race_unsafe",
            "main_inserted": 0, "plasma_inserted": 0,
            "main_skipped": 0, "plasma_skipped": 0,
        }

    if record.metadata and record.metadata.get("all_processed"):
        log.info(f"skip {record.file_name} (sha already processed)")
        return {
            "status": "skipped", "reason": "already_processed",
            "main_inserted": 0, "plasma_inserted": 0,
            "main_skipped": 0, "plasma_skipped": 0,
        }

    main_inserted = 0
    plasma_inserted = 0
    main_skipped = 0    # Step 23
    plasma_skipped = 0  # Step 23

    try:
        wb = load_workbook(record.file_path, read_only=True, data_only=True)
        try:
            if "Main Process" in wb.sheetnames:
                main_inserted, main_skipped = _process_sheet(
                    wb["Main Process"], MAIN_COL_MAP,
                    _INSERT_MAIN_SQL, record.id, "sputter_runs_auto_main"
                )
                log.info(
                    f"sputter_auto Main Process: +{main_inserted} rows, "
                    f"{main_skipped} skipped"
                )
            else:
                log.warning(f"'Main Process' sheet not found in {record.file_name}")

            if "Plasma Cleaning" in wb.sheetnames:
                plasma_inserted, plasma_skipped = _process_sheet(
                    wb["Plasma Cleaning"], PLASMA_COL_MAP,
                    _INSERT_PLASMA_SQL, record.id, "sputter_runs_auto_plasma"
                )
                log.info(
                    f"sputter_auto Plasma Cleaning: +{plasma_inserted} rows, "
                    f"{plasma_skipped} skipped"
                )
            else:
                log.warning(f"'Plasma Cleaning' sheet not found in {record.file_name}")
        finally:
            wb.close()

        new_metadata = {
            "all_processed": True,
            "main_inserted": main_inserted,
            "plasma_inserted": plasma_inserted,
            "main_skipped": main_skipped,
            "plasma_skipped": plasma_skipped,
        }
        with session_scope_writer() as s:
            s.execute(_UPDATE_METADATA_SQL, {
                "id": record.id,
                "metadata": json.dumps(new_metadata, ensure_ascii=False),
                "row_count": main_inserted + plasma_inserted,
                "parser_status": "ok",
                "parser_error": None,
            })

        log.info(
            f"sputter_auto {record.file_name}: "
            f"main +{main_inserted} ({main_skipped} skipped), "
            f"plasma +{plasma_inserted} ({plasma_skipped} skipped)"
        )
        return {
            "status": "ok",
            "main_inserted": main_inserted,
            "plasma_inserted": plasma_inserted,
            "main_skipped": main_skipped,
            "plasma_skipped": plasma_skipped,
        }

    except Exception as e:
        error_msg = str(e)
        log.error(f"sputter_auto {record.file_name} failed: {error_msg}", exc_info=True)
        with session_scope_writer() as s:
            s.execute(_UPDATE_METADATA_SQL, {
                "id": record.id,
                "metadata": json.dumps(record.metadata or {}, ensure_ascii=False),
                "row_count": main_inserted + plasma_inserted,
                "parser_status": "error",
                "parser_error": error_msg[:1000],
            })
        return {
            "status": "error", "error": error_msg,
            "main_inserted": main_inserted, "plasma_inserted": plasma_inserted,
            "main_skipped": main_skipped, "plasma_skipped": plasma_skipped,
        }
