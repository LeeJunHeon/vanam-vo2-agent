"""CH1.xlsx 파싱 — Main Process → sputter_runs, Plasma Cleaning → equipment_events.

xlsx는 메인 source (정확도 높음). Main Process 시트는 성공한 공정만 → status='SUCCESS'.

처리 정책:
- 시트별 row_count는 source_files.metadata JSONB에 저장:
  {"main_process_rows": N, "plasma_cleaning_rows": M}
  N/M은 이전 tick까지 처리한 row 수. 다음 tick에 N/M 이후 row만 처리.
- 첫 3 row는 header (section title / column name / unit), 4번째부터 데이터.
- xlsx 마지막에 빈 row 많음 → r[0] is not None 으로 필터.
- openpyxl read_only=True (메모리/잠금 안전). 외부 PC가 잡고 있어도 read-only로 열림.
- is_race_unsafe 면 처리 skip.

멱등성:
- sputter_runs: (source_file_id, row_number) UNIQUE → ON CONFLICT DO NOTHING
- equipment_events: UNIQUE 없음 → metadata 기반 prev_clean으로 중복 방지
  (단, sha 변경 시 새 source_file_id가 되어 prev_clean=0부터 시작 → 중복 위험.
   Phase 1b 통합 검증에서 발견되면 scan_files lookup으로 fix 예정)
"""
import json
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import text

from shared.db import session_scope_writer
from etl_worker.jobs.scan_files import SourceFileRecord

log = logging.getLogger("etl.parsers.sputter_xlsx")

XLSX_MAIN_COLS = [
    '날짜', '담당자', 'Process Name', '비고', '기판',
    'Main Shutter', 'Power Select',
    'G1 Target', 'G2 Target', 'G3 Target',
    'Dep.rate', 'Thickness', 'Chuck',
    'Shutter Delay', 'Process Time', 'Base Pressure',
    'SP Ar', 'Avg Ar', 'SP N2', 'Avg N2', 'SP O2', 'Avg O2',
    'SP Pressure', 'Avg Pressure',
    'Power Source', 'SP Power', 'Avg Power',
    'Avg For.p', 'Avg Ref.p', 'Avg Load', 'Avg Tune',
    'Avg Voltage', 'Avg Current',
    'Duty Cycle', 'Frequency', 'Off Time',
    'Soft Arc', 'Hard Arc',
]

XLSX_CLEANING_COLS = [
    '날짜', '담당자', 'Process Name', '비고', '기판',
    'Time', 'Base Pressure',
    'SP Ar', 'Avg Ar', 'SP Pressure', 'Avg Pressure',
    'SP Power', 'Avg For.p', 'Avg Ref.p',
    'Avg Load', 'Avg Tune',
]


def _serialize(v):
    """JSON 직렬화 헬퍼 — datetime → ISO string."""
    if isinstance(v, datetime):
        return v.isoformat()
    return v


def _normalize(v):
    """빈 문자열을 None으로. xlsx에서 빈 셀이 ''로 들어올 수 있음."""
    if v == '' or v is None:
        return None
    return v


_INSERT_MAIN_SQL = text("""
    INSERT INTO vo2.sputter_runs (
        sputter_run_id, source_file_id, row_number, chamber, sample_id,
        start_time, end_time, recipe_name, operator, note, substrate, status,
        target, power_source, power_select, main_shutter_open, chuck_position,
        shutter_delay_min, process_time_min, integration_time_s,
        sp_ar_sccm, sp_o2_sccm, sp_n2_sccm, sp_pressure_mtorr, sp_power_w,
        base_pressure_torr, avg_pressure_mtorr,
        avg_ar_sccm, avg_o2_sccm, avg_n2_sccm,
        avg_power_w, avg_for_p_w, avg_ref_p_w, avg_voltage_v, avg_current_a,
        avg_load_au, avg_tune_au,
        pulse_freq_khz, pulse_duty_cycle_pct, pulse_off_time_us,
        dep_rate_nm_per_s, thickness_nm, soft_arc_count, hard_arc_count,
        o2_ratio, total_flow_sccm, power_time_ws,
        substrate_temperature_c, target_cleaning_flag,
        raw_json
    )
    VALUES (
        :sputter_run_id, :source_file_id, :row_number, :chamber, :sample_id,
        :start_time, :end_time, :recipe_name, :operator, :note, :substrate, :status,
        :target, :power_source, :power_select, :main_shutter_open, :chuck_position,
        :shutter_delay_min, :process_time_min, :integration_time_s,
        :sp_ar_sccm, :sp_o2_sccm, :sp_n2_sccm, :sp_pressure_mtorr, :sp_power_w,
        :base_pressure_torr, :avg_pressure_mtorr,
        :avg_ar_sccm, :avg_o2_sccm, :avg_n2_sccm,
        :avg_power_w, :avg_for_p_w, :avg_ref_p_w, :avg_voltage_v, :avg_current_a,
        :avg_load_au, :avg_tune_au,
        :pulse_freq_khz, :pulse_duty_cycle_pct, :pulse_off_time_us,
        :dep_rate_nm_per_s, :thickness_nm, :soft_arc_count, :hard_arc_count,
        :o2_ratio, :total_flow_sccm, :power_time_ws,
        :substrate_temperature_c, :target_cleaning_flag,
        CAST(:raw_json AS JSONB)
    )
    ON CONFLICT (source_file_id, row_number) DO NOTHING
""")

_INSERT_CLEANING_SQL = text("""
    INSERT INTO vo2.equipment_events (
        source_file_id, event_time, equipment, chamber, event_type,
        operator, detail, source_kind, raw_json
    )
    VALUES (
        :source_file_id, :event_time, :equipment, :chamber, :event_type,
        :operator, :detail, :source_kind, CAST(:raw_json AS JSONB)
    )
""")

_UPDATE_METADATA_SQL = text("""
    UPDATE vo2.source_files
    SET metadata = CAST(:metadata AS JSONB),
        row_count = :row_count,
        parser_status = :parser_status,
        parser_error = :parser_error,
        last_indexed_at = NOW()
    WHERE id = :id
""")


def _build_main_payload(row_dict: dict, source_file_id: int, row_number: int) -> dict:
    """xlsx Main Process row → sputter_runs INSERT payload."""
    ar = _normalize(row_dict.get('Avg Ar'))
    o2 = _normalize(row_dict.get('Avg O2'))
    n2 = _normalize(row_dict.get('Avg N2'))
    avg_power = _normalize(row_dict.get('Avg Power'))
    process_time_min = _normalize(row_dict.get('Process Time'))
    start_time = row_dict.get('날짜')

    # 파생값
    o2_ratio = None
    if isinstance(ar, (int, float)) and isinstance(o2, (int, float)) and (ar + o2) > 0:
        o2_ratio = o2 / (ar + o2)
    total_flow = None
    flows = [x for x in (ar, o2, n2) if isinstance(x, (int, float))]
    if flows:
        total_flow = sum(flows)
    power_time = None
    if isinstance(avg_power, (int, float)) and isinstance(process_time_min, (int, float)):
        power_time = avg_power * process_time_min * 60.0

    if isinstance(start_time, datetime):
        sputter_run_id = f"CH1-{start_time.strftime('%Y%m%d-%H%M%S')}"
    else:
        sputter_run_id = f"CH1-row-{source_file_id}-{row_number}"

    return {
        'sputter_run_id':         sputter_run_id,
        'source_file_id':         source_file_id,
        'row_number':             row_number,
        'chamber':                'CH1',
        'sample_id':              None,
        'start_time':             start_time if isinstance(start_time, datetime) else None,
        'end_time':               None,
        'recipe_name':            _normalize(row_dict.get('Process Name')),
        'operator':               _normalize(row_dict.get('담당자')),
        'note':                   _normalize(row_dict.get('비고')),
        'substrate':              _normalize(row_dict.get('기판')),
        'status':                 'SUCCESS',
        'target':                 _normalize(row_dict.get('G1 Target')),
        'power_source':           _normalize(row_dict.get('Power Source')),
        'power_select':           _normalize(row_dict.get('Power Select')),
        'main_shutter_open':      row_dict.get('Main Shutter') == 'T',
        'chuck_position':         _normalize(row_dict.get('Chuck')),
        'shutter_delay_min':      _normalize(row_dict.get('Shutter Delay')),
        'process_time_min':       process_time_min,
        'integration_time_s':     None,
        'sp_ar_sccm':             _normalize(row_dict.get('SP Ar')),
        'sp_o2_sccm':             _normalize(row_dict.get('SP O2')),
        'sp_n2_sccm':             _normalize(row_dict.get('SP N2')),
        'sp_pressure_mtorr':      _normalize(row_dict.get('SP Pressure')),
        'sp_power_w':             _normalize(row_dict.get('SP Power')),
        'base_pressure_torr':     _normalize(row_dict.get('Base Pressure')),
        'avg_pressure_mtorr':     _normalize(row_dict.get('Avg Pressure')),
        'avg_ar_sccm':            ar,
        'avg_o2_sccm':            o2,
        'avg_n2_sccm':            n2,
        'avg_power_w':            avg_power,
        'avg_for_p_w':            _normalize(row_dict.get('Avg For.p')),
        'avg_ref_p_w':            _normalize(row_dict.get('Avg Ref.p')),
        'avg_voltage_v':          _normalize(row_dict.get('Avg Voltage')),
        'avg_current_a':          _normalize(row_dict.get('Avg Current')),
        'avg_load_au':            _normalize(row_dict.get('Avg Load')),
        'avg_tune_au':            _normalize(row_dict.get('Avg Tune')),
        'pulse_freq_khz':         _normalize(row_dict.get('Frequency')),
        'pulse_duty_cycle_pct':   _normalize(row_dict.get('Duty Cycle')),
        'pulse_off_time_us':      _normalize(row_dict.get('Off Time')),
        'dep_rate_nm_per_s':      _normalize(row_dict.get('Dep.rate')),
        'thickness_nm':           _normalize(row_dict.get('Thickness')),
        'soft_arc_count':         _normalize(row_dict.get('Soft Arc')),
        'hard_arc_count':         _normalize(row_dict.get('Hard Arc')),
        'o2_ratio':               o2_ratio,
        'total_flow_sccm':        total_flow,
        'power_time_ws':          power_time,
        'substrate_temperature_c': None,
        'target_cleaning_flag':   False,
        'raw_json':               json.dumps(
            {k: _serialize(v) for k, v in row_dict.items()}, ensure_ascii=False
        ),
    }


def _build_cleaning_payload(row_dict: dict, source_file_id: int) -> dict:
    """xlsx Plasma Cleaning row → equipment_events INSERT payload."""
    parts = []
    if (t := _normalize(row_dict.get('Time'))) is not None:
        parts.append(f"Time={t}min")
    if (p := _normalize(row_dict.get('Avg For.p'))) is not None:
        parts.append(f"Power={p}W")
    if (a := _normalize(row_dict.get('Avg Ar'))) is not None:
        parts.append(f"Ar={a}sccm")
    detail = ", ".join(parts) if parts else None

    event_time = row_dict.get('날짜')
    return {
        'source_file_id': source_file_id,
        'event_time':     event_time if isinstance(event_time, datetime) else None,
        'equipment':      'ch1',
        'chamber':        'CH1',
        'event_type':     'plasma_cleaning',
        'operator':       _normalize(row_dict.get('담당자')),
        'detail':         detail,
        'source_kind':    'auto',
        'raw_json':       json.dumps(
            {k: _serialize(v) for k, v in row_dict.items()}, ensure_ascii=False
        ),
    }


def _read_sheet_rows(file_path: Path, sheet_name: str, columns: list) -> list[dict]:
    """xlsx 한 시트를 읽어 dict list 반환. 첫 3 row(header) 제외, 빈 row 제외."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            log.warning(f"sheet '{sheet_name}' not found in {file_path.name}")
            return []
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    data_rows = [r for r in rows[3:] if r and r[0] is not None]
    return [dict(zip(columns, r)) for r in data_rows]


def parse_xlsx(record: SourceFileRecord) -> dict:
    """xlsx 두 시트 모두 처리 + source_files 메타데이터 갱신.

    Returns: {"main_inserted": N, "cleaning_inserted": M, "status": "ok"|"error"|"skipped", ...}
    """
    if record.is_race_unsafe:
        log.info(f"skip {record.file_name} (race_unsafe, mtime too recent)")
        return {"main_inserted": 0, "cleaning_inserted": 0, "status": "skipped"}

    prev = record.metadata or {}
    prev_main = int(prev.get("main_process_rows", 0))
    prev_clean = int(prev.get("plasma_cleaning_rows", 0))

    main_inserted = 0
    cleaning_inserted = 0

    try:
        main_rows = _read_sheet_rows(record.file_path, 'Main Process', XLSX_MAIN_COLS)
        new_main = main_rows[prev_main:]
        clean_rows = _read_sheet_rows(record.file_path, 'Plasma Cleaning', XLSX_CLEANING_COLS)
        new_clean = clean_rows[prev_clean:]

        with session_scope_writer() as s:
            for offset, row_dict in enumerate(new_main):
                payload = _build_main_payload(row_dict, record.id, prev_main + offset)
                s.execute(_INSERT_MAIN_SQL, payload)
                main_inserted += 1

            for row_dict in new_clean:
                payload = _build_cleaning_payload(row_dict, record.id)
                s.execute(_INSERT_CLEANING_SQL, payload)
                cleaning_inserted += 1

            new_metadata = {
                "main_process_rows":    prev_main + main_inserted,
                "plasma_cleaning_rows": prev_clean + cleaning_inserted,
            }
            s.execute(_UPDATE_METADATA_SQL, {
                "id": record.id,
                "metadata": json.dumps(new_metadata, ensure_ascii=False),
                "row_count": prev_main + main_inserted,
                "parser_status": "ok",
                "parser_error": None,
            })

        log.info(
            f"xlsx {record.file_name}: main +{main_inserted} (total {prev_main + main_inserted}), "
            f"cleaning +{cleaning_inserted} (total {prev_clean + cleaning_inserted})"
        )
        return {
            "main_inserted":     main_inserted,
            "cleaning_inserted": cleaning_inserted,
            "status":            "ok",
        }

    except Exception as e:
        error_msg = str(e)
        log.error(f"xlsx {record.file_name} parse failed: {error_msg}", exc_info=True)
        with session_scope_writer() as s:
            s.execute(_UPDATE_METADATA_SQL, {
                "id": record.id,
                "metadata": json.dumps(record.metadata or {}, ensure_ascii=False),
                "row_count": prev_main,
                "parser_status": "error",
                "parser_error": error_msg[:1000],
            })
        return {
            "main_inserted":     main_inserted,
            "cleaning_inserted": cleaning_inserted,
            "status":            "error",
            "error":             error_msg,
        }
