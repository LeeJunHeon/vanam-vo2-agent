"""사람 sputter log xlsx 파서 — Phase 4 Step 22: forward fill + 분류.

전략 변경 (Step 22):
- 시트의 '날짜' 컬럼이 5가지 의미로 쓰임 → 분류 + forward fill
- 모든 데이터 보존 + process_date 정확성 100%

xlsx 구조:
- '통합' 시트: R1=헤더, R2=단위, R3부터 데이터
- C0~C22 (W까지) 23 컬럼 매핑. C23+ 무시.

'날짜' 컬럼 분류 (raw_date_type):
- yyyymmdd_pure_marker: YYYYMMDD 큰 숫자 + 우측 비어있음 (담당='Pre' 또는 빈칸) → row_kind='marker'
- yyyymmdd_integrated: YYYYMMDD 큰 숫자 + 우측 데이터 있음 (옛 패턴) → row_kind='process'
- datetime_as_marker: 2025-06-09 이전 datetime → 진짜 마커 (시기 1, 옛 옛 운영 패턴)
- datetime_excel: 2025-06-09 이후 datetime → Excel 자동변환 의심, forward fill
- small_number: 1, 2, 3, ..., 7 같은 작은 숫자 (그 날의 N번째 공정 순서) → forward fill
- text: 'depo', 'depo1', 'RFP2', 'DCP 1' 등 텍스트 → forward fill
- null_with_data: '날짜' 빈칸 + 다른 컬럼 데이터 있음 → forward fill
- fully_empty: 모든 컬럼 빈 row → row_kind='empty'
- orphan_before_first_marker: 첫 마커 만나기 전 row → process_date NULL

forward fill:
- 마커 row 만나면 current_date 갱신
- 마커 아닌 row의 process_date = current_date

빈 row 판정:
- 23 컬럼 모두 비어있어야 fully_empty.
- 연속 5개 fully_empty면 데이터 끝 (break).
"""

import json
import logging
from datetime import datetime, date

from openpyxl import load_workbook
from sqlalchemy import text

from shared.db import session_scope_writer
from etl_worker.jobs.scan_files import SourceFileRecord

log = logging.getLogger("etl.parsers.sputter_human")

SHEET_NAME = "통합"

# YYYYMMDD 마커 판정 범위
YYYYMMDD_MIN = 20240101
YYYYMMDD_MAX = 20271231

# Phase 4 Step 22-C: 시기 1 datetime을 마커로 인식하는 기준일
# - 이 날짜 이전 datetime row → 진짜 진행 날짜 마커 (옛 옛 운영 패턴)
# - 이 날짜 이후 datetime row → Excel 자동변환 의심 (운영자가 '6-1' 입력 시 Excel이 자동변환), forward fill만
DATETIME_AS_MARKER_BEFORE = date(2025, 6, 9)

# (col_idx, db_col, type) type: 'real' | 'text'
# process_date는 forward fill로 별도 처리하므로 여기서 제외
HUMAN_COL_MAP = [
    (1, 'operator', 'text'),
    (2, 'sub_label_raw', 'text'),
    (3, 'pc_gas', 'text'),
    (4, 'pc_power_w', 'real'),
    (5, 'pc_pressure_mtorr', 'real'),
    (6, 'pc_gas_flow_sccm', 'real'),
    (7, 'pc_time_min', 'real'),
    (8, 'shutter_delay_min', 'real'),
    (9, 'sp_power_w', 'real'),
    (10, 'sp_flow_sccm', 'real'),
    (11, 'sp_pressure_mtorr', 'real'),
    (12, 'sp_time', 'text'),
    (13, 'thickness', 'text'),
    (14, 'furnace_type', 'text'),
    (15, 'annealing_temp', 'text'),
    (16, 'annealing_gas_flow', 'text'),
    (17, 'pulsed_dc_freq', 'text'),
    (18, 'off_time_us', 'real'),
    (19, 'duty', 'text'),
    (20, 'depo_rate', 'text'),
    (21, 'target', 'text'),
    (22, 'notes', 'text'),
]


def _to_real(v):
    if v is None or v == '':
        return None
    if isinstance(v, str) and not v.strip():
        return None
    try:
        f = float(v)
        if f != f or f == float('inf') or f == float('-inf'):
            return None
        return f
    except (ValueError, TypeError):
        return None


def _to_text(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return str(v)


def _serialize(v):
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return v


def _row_to_full_dict(row, header):
    result = {}
    for i, val in enumerate(row):
        key = header[i] if i < len(header) and header[i] else f'col_{i}'
        if isinstance(key, str):
            key = key.replace('\n', ' ').strip()
        result[key] = val
    return result


def _is_fully_empty(row, max_col=23):
    for i in range(max_col):
        if i < len(row):
            v = row[i]
            if v is not None and not (isinstance(v, str) and not v.strip()):
                return False
    return True


def _count_right_data(row, start_col=1, end_col=23):
    count = 0
    for i in range(start_col, end_col):
        if i < len(row):
            v = row[i]
            if v is not None and not (isinstance(v, str) and not v.strip()):
                count += 1
    return count


def _classify_date_cell(date_val, operator, right_count):
    """returns (raw_date_type, marker_yyyymmdd_or_None, row_kind, process_seq_or_None)."""
    if isinstance(date_val, (int, float)):
        v = int(date_val)
        if YYYYMMDD_MIN <= v <= YYYYMMDD_MAX:
            is_pure_marker = (
                operator == "Pre"
                or (operator is None and right_count <= 1)
                or right_count <= 2
            )
            if is_pure_marker:
                return ('yyyymmdd_pure_marker', v, 'marker', None)
            else:
                return ('yyyymmdd_integrated', v, 'process', None)
        else:
            return ('small_number', None, 'process', v)

    if isinstance(date_val, datetime):
        d = date_val.date()
        if d < DATETIME_AS_MARKER_BEFORE:
            # 시기 1 — 진짜 진행 날짜 마커로 처리
            v = d.year * 10000 + d.month * 100 + d.day
            return ('datetime_as_marker', v, 'process', None)
        # 시기 2/3 — Excel 자동변환 의심 (forward fill만)
        return ('datetime_excel', None, 'process', None)
    if isinstance(date_val, date):
        if date_val < DATETIME_AS_MARKER_BEFORE:
            v = date_val.year * 10000 + date_val.month * 100 + date_val.day
            return ('datetime_as_marker', v, 'process', None)
        return ('datetime_excel', None, 'process', None)

    if isinstance(date_val, str):
        s = date_val.strip()
        if s:
            return ('text', None, 'process', None)
        date_val = None

    if date_val is None:
        if right_count == 0:
            return ('fully_empty', None, 'empty', None)
        else:
            return ('null_with_data', None, 'process', None)

    return ('text', None, 'process', None)


def _yyyymmdd_to_date(v):
    s = str(int(v))
    if len(s) != 8:
        return None
    try:
        y = int(s[0:4])
        m = int(s[4:6])
        d = int(s[6:8])
        return date(y, m, d)
    except (ValueError, TypeError):
        return None


_DB_COLS = ['process_date'] + [m[1] for m in HUMAN_COL_MAP] + [
    'raw_date_value', 'raw_date_type', 'process_seq_in_day', 'row_kind'
]
_INSERT_HUMAN_SQL = text(f"""
    INSERT INTO vo2.sputter_runs_human (
        {', '.join(_DB_COLS)},
        source_file_id, row_number, raw_json, parse_status
    ) VALUES (
        {', '.join(':' + c for c in _DB_COLS)},
        :source_file_id, :row_number, CAST(:raw_json AS JSONB), :parse_status
    )
    ON CONFLICT (source_file_id, row_number) DO NOTHING
""")

_UPDATE_METADATA_SQL = text("""
    UPDATE vo2.source_files
    SET metadata = CAST(:metadata AS JSONB),
        row_count = :row_count,
        parser_status = :parser_status,
        parser_error = :parser_error,
        last_indexed_at = NOW()
    WHERE id = :id
""")


def _build_payload(
    row, source_file_id, row_number, header,
    current_date, raw_date_type, row_kind, process_seq
):
    payload = {}

    if current_date is not None:
        payload['process_date'] = datetime.combine(current_date, datetime.min.time())
    else:
        payload['process_date'] = None

    for col_idx, db_col, dtype in HUMAN_COL_MAP:
        v = row[col_idx] if col_idx < len(row) else None
        if dtype == 'real':
            payload[db_col] = _to_real(v)
        else:
            payload[db_col] = _to_text(v)

    raw_date_val = row[0] if len(row) > 0 else None
    if raw_date_val is None:
        payload['raw_date_value'] = None
    elif isinstance(raw_date_val, (datetime, date)):
        payload['raw_date_value'] = raw_date_val.isoformat()
    else:
        payload['raw_date_value'] = str(raw_date_val)

    payload['raw_date_type'] = raw_date_type
    payload['process_seq_in_day'] = process_seq
    payload['row_kind'] = row_kind

    payload['source_file_id'] = source_file_id
    payload['row_number'] = row_number
    payload['parse_status'] = None

    raw = _row_to_full_dict(row, header)
    payload['raw_json'] = json.dumps(
        {k: _serialize(v) for k, v in raw.items()},
        ensure_ascii=False,
    )
    return payload


def parse_sputter_human(record: SourceFileRecord) -> dict:
    if record.is_race_unsafe:
        log.info(f"skip {record.file_name} (race_unsafe)")
        return {"status": "skipped", "reason": "race_unsafe", "inserted": 0}

    if record.metadata and record.metadata.get("all_processed"):
        log.info(f"skip {record.file_name} (sha already processed)")
        return {"status": "skipped", "reason": "already_processed", "inserted": 0}

    inserted = 0
    marker_count = 0
    empty_count = 0

    try:
        wb = load_workbook(record.file_path, read_only=True, data_only=True)
        try:
            if SHEET_NAME not in wb.sheetnames:
                log.error(f"sheet '{SHEET_NAME}' not found in {record.file_name}")
                with session_scope_writer() as s:
                    s.execute(_UPDATE_METADATA_SQL, {
                        "id": record.id,
                        "metadata": json.dumps(record.metadata or {}, ensure_ascii=False),
                        "row_count": 0,
                        "parser_status": "error",
                        "parser_error": f"sheet '{SHEET_NAME}' not found",
                    })
                return {"status": "error", "error": f"sheet '{SHEET_NAME}' not found", "inserted": 0}

            ws = wb[SHEET_NAME]

            header_data = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            header = list(header_data[0]) if header_data else []
            header = [
                (h.replace('\n', ' ').strip() if isinstance(h, str) else h)
                for h in header
            ]

            current_date = None
            empty_streak = 0

            with session_scope_writer() as s:
                for row_idx, row in enumerate(
                    ws.iter_rows(min_row=3, values_only=True), start=3
                ):
                    if _is_fully_empty(row, max_col=23):
                        empty_streak += 1
                        if empty_streak >= 5:
                            break
                        payload = _build_payload(
                            row, record.id, row_idx, header,
                            current_date, 'fully_empty', 'empty', None
                        )
                        s.execute(_INSERT_HUMAN_SQL, payload)
                        empty_count += 1
                        inserted += 1
                        continue
                    empty_streak = 0

                    date_val = row[0] if len(row) > 0 else None
                    operator = row[1] if len(row) > 1 else None
                    if isinstance(operator, str):
                        operator = operator.strip()
                    right_count = _count_right_data(row, start_col=1, end_col=23)

                    raw_type, marker_yyyymmdd, row_kind, process_seq = _classify_date_cell(
                        date_val, operator, right_count
                    )

                    if marker_yyyymmdd is not None:
                        new_date = _yyyymmdd_to_date(marker_yyyymmdd)
                        if new_date is not None:
                            current_date = new_date

                    if row_kind == 'process' and current_date is None:
                        raw_type = 'orphan_before_first_marker'

                    payload = _build_payload(
                        row, record.id, row_idx, header,
                        current_date, raw_type, row_kind, process_seq
                    )
                    s.execute(_INSERT_HUMAN_SQL, payload)
                    inserted += 1

                    if row_kind == 'marker':
                        marker_count += 1

        finally:
            wb.close()

        new_metadata = {
            "all_processed": True,
            "inserted": inserted,
            "marker_rows": marker_count,
            "empty_rows": empty_count,
        }
        with session_scope_writer() as s:
            s.execute(_UPDATE_METADATA_SQL, {
                "id": record.id,
                "metadata": json.dumps(new_metadata, ensure_ascii=False),
                "row_count": inserted,
                "parser_status": "ok",
                "parser_error": None,
            })

        log.info(
            f"sputter_human {record.file_name}: +{inserted} rows "
            f"(markers={marker_count}, empty={empty_count})"
        )
        return {
            "status": "ok",
            "inserted": inserted,
            "marker_rows": marker_count,
            "empty_rows": empty_count,
        }

    except Exception as e:
        error_msg = str(e)
        log.error(
            f"sputter_human {record.file_name} parse failed: {error_msg}",
            exc_info=True,
        )
        with session_scope_writer() as s:
            s.execute(_UPDATE_METADATA_SQL, {
                "id": record.id,
                "metadata": json.dumps(record.metadata or {}, ensure_ascii=False),
                "row_count": inserted,
                "parser_status": "error",
                "parser_error": error_msg[:1000],
            })
        return {"status": "error", "error": error_msg, "inserted": inserted}
