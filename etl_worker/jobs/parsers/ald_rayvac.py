"""ALD Rayvac xlsx 파서 — 단일 시트 → ald_rayvac_runs.

Rayvac xlsx 구조:
- '공정 레시피 & 결과 정리' 시트: R1=컬럼명, R2부터 데이터, 39 컬럼
- chemistry 분리 없음 (Rayvac은 TTIP만, TDMAT 없음)
- Oxidant는 O3 (NCD는 H2O)
- 첫 컬럼('공정 번호')이 비면 → 데이터 끝

처리 정책 (NCD와 동일):
- 새 sha (metadata 비어있음) → 시트 처리
- 같은 sha (metadata에 'all_processed') → skip
- xlsx 변경 시 새 source_file_id로 새 row INSERT, 옛 데이터 그대로 보존
- 검증 실패 row → parse_errors 격리 (GPT 친화 자연어 메시지)

멱등성:
- ald_rayvac_runs UNIQUE (source_file_id, row_number) → ON CONFLICT DO NOTHING
- parse_errors UNIQUE (source_file_id, row_number, error_type) → ON CONFLICT DO NOTHING
"""

import json
import logging
from datetime import datetime, date
from typing import Optional

from openpyxl import load_workbook
from sqlalchemy import text

from shared.db import session_scope_writer
from etl_worker.jobs.scan_files import SourceFileRecord

log = logging.getLogger("etl.parsers.ald_rayvac")

RAYVAC_SHEET = "공정 레시피 & 결과 정리"

# Rayvac R1 헤더 (0-indexed 컬럼 위치 → init.sql 컬럼명)
# C1 '온도/습도', C24 'plasma cleaning', C28 'Virtual Max-min(%)',
# C29~ '특이 사항', '장비 이슈', 'Base(net)', '예측' 등은 raw_json만 보존
RAYVAC_COL_MAP = {
    0:  'batch_no_raw',           # 공정 번호
    2:  'process_date_raw',       # Date
    3:  'stable_time_min',        # Stable time (min)
    4:  'stage_temp_c',           # Stage Temp (°C)
    5:  'body_temp_c',            # Body Temp (°C)
    6:  'top_temp_c',             # Top Temp (°C)
    7:  'stage_height_mm',        # Stage Height (mm)
    8:  'precursor_line_temp_c',  # Precursor line Temp (°C)
    9:  'reactant_line_temp_c',   # Reactant line Temp (°C)
    10: 'base_pressure_torr',     # Base pressure [Torr]
    11: 'throttle_pct',           # Throttle [%]
    12: 'ttip_temp_c',            # TTIP Temp [°C]
    13: 'source_base_sccm',       # Source Base [sccm]
    14: 'ttip_assist_sccm',       # TTIP Assist [sccm]
    15: 'reactant_base_sccm',     # Reactant Base [sccm]
    16: 'o3_conc',                # O3 conc
    17: 'o2_flow_sccm',           # O2 Flow (sccm)
    18: 'ttip_assist_time_s',     # TTIP Assist time (sec)
    19: 'ttip_pulse_s',           # TTIP Pulse time (sec)
    20: 'ttip_purge_s',           # TTIP Purge time (sec)
    21: 'o3_pulse_s',             # O3 Pulse time (sec)
    22: 'o3_purge_s',             # O3 Purge time (sec)
    23: 'cycles',                 # Cycle (#)
    24: 'plasma_cleaning_flag',   # plasma cleaning (truthy → True)
    25: 'gpc_a_per_cycle',        # GPC (A/cycle)
    26: 'max_min_pct',            # Max-min(%)
    27: 'virtual_max_min_pct',    # Virtual Max-min(%)
}


def _to_float(v):
    """REAL 컬럼용 안전 변환."""
    if v is None or v == '':
        return None
    try:
        f = float(v)
        if f != f or f == float('inf') or f == float('-inf'):
            return None
        return f
    except (ValueError, TypeError):
        return None


def _to_int(v):
    """INTEGER 컬럼용 안전 변환."""
    if v is None or v == '':
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def _to_date(v):
    """DATE 컬럼용 안전 변환."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _to_bool(v):
    """BOOLEAN 컬럼용 안전 변환. None/빈 문자열 → None.

    Rayvac 'plasma cleaning' 컬럼은 임의 값일 수 있어 truthy 판정.
    'O', 'X', 'T', 'F', '있음', '없음', 1, 0 등 다양한 케이스 대응.
    """
    if v is None or v == '':
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(v)
    if isinstance(v, str):
        s = v.strip().lower()
        if s in ('o', 't', 'true', 'y', 'yes', '있음', 'on', '1'):
            return True
        if s in ('x', 'f', 'false', 'n', 'no', '없음', 'off', '0'):
            return False
        # 그 외 비어있지 않은 문자열은 truthy로 판정 (예: '있음 (1회)')
        return True
    return None


def _serialize(v):
    """JSON 직렬화 헬퍼 — datetime/date → ISO string."""
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return v


def _row_to_full_dict(row: tuple, header: list) -> dict:
    """xlsx row tuple → 전체 dict (raw_json 보존용)."""
    result = {}
    for i, val in enumerate(row):
        key = header[i] if i < len(header) and header[i] else f'col_{i}'
        if isinstance(key, str):
            key = key.replace('\n', ' ').strip()
        result[key] = val
    return result


_INSERT_RAYVAC_SQL = text("""
    INSERT INTO vo2.ald_rayvac_runs (
        batch_no, process_date,
        stable_time_min, stage_temp_c, body_temp_c, top_temp_c, stage_height_mm,
        precursor_line_temp_c, reactant_line_temp_c,
        base_pressure_torr, throttle_pct,
        ttip_temp_c, source_base_sccm, ttip_assist_sccm, reactant_base_sccm,
        o3_conc, o2_flow_sccm,
        ttip_assist_time_s, ttip_pulse_s, ttip_purge_s,
        o3_pulse_s, o3_purge_s,
        cycles, plasma_cleaning_flag,
        gpc_a_per_cycle, max_min_pct, virtual_max_min_pct,
        source_file_id, row_number, raw_json, parse_status
    ) VALUES (
        :batch_no, :process_date,
        :stable_time_min, :stage_temp_c, :body_temp_c, :top_temp_c, :stage_height_mm,
        :precursor_line_temp_c, :reactant_line_temp_c,
        :base_pressure_torr, :throttle_pct,
        :ttip_temp_c, :source_base_sccm, :ttip_assist_sccm, :reactant_base_sccm,
        :o3_conc, :o2_flow_sccm,
        :ttip_assist_time_s, :ttip_pulse_s, :ttip_purge_s,
        :o3_pulse_s, :o3_purge_s,
        :cycles, :plasma_cleaning_flag,
        :gpc_a_per_cycle, :max_min_pct, :virtual_max_min_pct,
        :source_file_id, :row_number, CAST(:raw_json AS JSONB), :parse_status
    )
    ON CONFLICT (source_file_id, row_number) DO NOTHING
""")

_INSERT_PARSE_ERROR_SQL = text("""
    INSERT INTO vo2.parse_errors (
        source_file_id, row_number, error_type, error_detail, raw_data
    ) VALUES (
        :source_file_id, :row_number, :error_type, :error_detail,
        CAST(:raw_data AS JSONB)
    )
    ON CONFLICT (source_file_id, row_number, error_type) DO NOTHING
""")

_UPDATE_METADATA_SQL = text("""
    UPDATE vo2.source_files
    SET metadata = CAST(:metadata AS JSONB),
        row_count = :row_count,
        parser_status = :parser_status,
        parser_error = :parser_error,
        last_indexed_at = NOW()
    WHERE id = :id
""")


def _validate_row(row: tuple, row_idx: int) -> Optional[dict]:
    """row 검증. 격리할 에러가 있으면 dict 반환."""
    batch_raw = row[0] if len(row) > 0 else None
    date_raw = row[2] if len(row) > 2 else None  # Rayvac은 C2가 Date (NCD는 C1)

    # 1. batch_no 정수 변환 검증
    try:
        int(batch_raw) if batch_raw is not None else None
        if batch_raw is None:
            return None
    except (ValueError, TypeError):
        return {
            "error_type": "batch_no_invalid",
            "error_detail": (
                f"Rayvac '{RAYVAC_SHEET}' 시트 R{row_idx} row의 공정 번호가 "
                f"'{batch_raw}'로 정수가 아닙니다. ald_rayvac_runs에 INSERT하지 못했습니다. "
                f"운영자가 xlsx의 R{row_idx} 첫 컬럼을 정확한 정수로 수정하면 "
                f"다음 5분 tick에서 자동 처리됩니다 (단 이 row는 새 source_file_id로 들어가니 "
                f"옛 source_file_id의 격리 row는 그대로 남습니다)."
            ),
        }

    # 2. Date 검증
    if not isinstance(date_raw, (datetime, date)):
        return {
            "error_type": "date_missing",
            "error_detail": (
                f"Rayvac '{RAYVAC_SHEET}' 시트 R{row_idx} row의 Date 컬럼이 비어있거나 "
                f"datetime이 아닙니다 (실제 값: {date_raw!r}). "
                f"ald_rayvac_runs.process_date는 NOT NULL이라 INSERT하지 못했습니다. "
                f"운영자가 xlsx에 Date를 입력하면 다음 tick에서 자동 처리됩니다."
            ),
        }

    return None


def _build_payload(
    row: tuple, source_file_id: int, row_number: int, header: list
) -> dict:
    """검증 통과한 row → ald_rayvac_runs INSERT payload."""
    payload = {
        'batch_no': _to_int(row[0]),
        'process_date': _to_date(row[2]),
        'stable_time_min': _to_float(row[3]) if len(row) > 3 else None,
        'stage_temp_c': _to_float(row[4]) if len(row) > 4 else None,
        'body_temp_c': _to_float(row[5]) if len(row) > 5 else None,
        'top_temp_c': _to_float(row[6]) if len(row) > 6 else None,
        'stage_height_mm': _to_float(row[7]) if len(row) > 7 else None,
        'precursor_line_temp_c': _to_float(row[8]) if len(row) > 8 else None,
        'reactant_line_temp_c': _to_float(row[9]) if len(row) > 9 else None,
        'base_pressure_torr': _to_float(row[10]) if len(row) > 10 else None,
        'throttle_pct': _to_float(row[11]) if len(row) > 11 else None,
        'ttip_temp_c': _to_float(row[12]) if len(row) > 12 else None,
        'source_base_sccm': _to_float(row[13]) if len(row) > 13 else None,
        'ttip_assist_sccm': _to_float(row[14]) if len(row) > 14 else None,
        'reactant_base_sccm': _to_float(row[15]) if len(row) > 15 else None,
        'o3_conc': _to_float(row[16]) if len(row) > 16 else None,
        'o2_flow_sccm': _to_float(row[17]) if len(row) > 17 else None,
        'ttip_assist_time_s': _to_float(row[18]) if len(row) > 18 else None,
        'ttip_pulse_s': _to_float(row[19]) if len(row) > 19 else None,
        'ttip_purge_s': _to_float(row[20]) if len(row) > 20 else None,
        'o3_pulse_s': _to_float(row[21]) if len(row) > 21 else None,
        'o3_purge_s': _to_float(row[22]) if len(row) > 22 else None,
        'cycles': _to_int(row[23]) if len(row) > 23 else None,
        'plasma_cleaning_flag': _to_bool(row[24]) if len(row) > 24 else None,
        'gpc_a_per_cycle': _to_float(row[25]) if len(row) > 25 else None,
        'max_min_pct': _to_float(row[26]) if len(row) > 26 else None,
        'virtual_max_min_pct': _to_float(row[27]) if len(row) > 27 else None,
        'source_file_id': source_file_id,
        'row_number': row_number,
        'parse_status': None,
    }

    raw = _row_to_full_dict(row, header)
    payload['raw_json'] = json.dumps(
        {k: _serialize(v) for k, v in raw.items()},
        ensure_ascii=False,
    )
    return payload


def _process_sheet(ws, source_file_id: int) -> tuple[int, int]:
    """Rayvac 단일 시트 처리. (inserted, errors) 반환."""
    inserted = 0
    errors = 0

    # R1 헤더 (Rayvac은 R1이 컬럼명, NCD R2와 다름)
    header_rows = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header = list(header_rows[0]) if header_rows else []
    header = [
        (h.replace('\n', ' ').strip() if isinstance(h, str) else h)
        for h in header
    ]

    with session_scope_writer() as s:
        # Rayvac은 R2부터 데이터
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            first_cell = row[0] if row else None

            if first_cell is None or (
                isinstance(first_cell, str) and not first_cell.strip()
            ):
                break

            error = _validate_row(row, row_idx)
            if error:
                raw_data = _row_to_full_dict(row, header)
                s.execute(_INSERT_PARSE_ERROR_SQL, {
                    "source_file_id": source_file_id,
                    "row_number": row_idx,
                    "error_type": error["error_type"],
                    "error_detail": error["error_detail"],
                    "raw_data": json.dumps(
                        {k: _serialize(v) for k, v in raw_data.items()},
                        ensure_ascii=False,
                    ),
                })
                errors += 1
                continue

            payload = _build_payload(row, source_file_id, row_idx, header)
            s.execute(_INSERT_RAYVAC_SQL, payload)
            inserted += 1

    log.info(f"Rayvac 시트 처리: +{inserted} rows, {errors} errors")
    return inserted, errors


def parse_ald_rayvac(record: SourceFileRecord) -> dict:
    """ALD Rayvac xlsx 메인 진입.

    Returns:
        {"status": "ok"|"error"|"skipped", "inserted": N, "errors": M, ...}
    """
    if record.is_race_unsafe:
        log.info(f"skip {record.file_name} (race_unsafe, mtime too recent)")
        return {
            "status": "skipped", "reason": "race_unsafe",
            "inserted": 0, "errors": 0,
        }

    if record.metadata and record.metadata.get("all_processed"):
        log.info(f"skip {record.file_name} (sha already processed)")
        return {
            "status": "skipped", "reason": "already_processed",
            "inserted": 0, "errors": 0,
        }

    inserted = 0
    errors = 0

    try:
        wb = load_workbook(record.file_path, read_only=True, data_only=True)
        try:
            if RAYVAC_SHEET not in wb.sheetnames:
                log.error(
                    f"sheet '{RAYVAC_SHEET}' not found in {record.file_name}. "
                    f"available sheets: {wb.sheetnames}"
                )
                with session_scope_writer() as s:
                    s.execute(_UPDATE_METADATA_SQL, {
                        "id": record.id,
                        "metadata": json.dumps(
                            record.metadata or {}, ensure_ascii=False
                        ),
                        "row_count": 0,
                        "parser_status": "error",
                        "parser_error": f"sheet '{RAYVAC_SHEET}' not found",
                    })
                return {
                    "status": "error",
                    "error": f"sheet '{RAYVAC_SHEET}' not found",
                    "inserted": 0, "errors": 0,
                }

            ws = wb[RAYVAC_SHEET]
            inserted, errors = _process_sheet(ws, record.id)
        finally:
            wb.close()

        new_metadata = {
            "all_processed": True,
            "inserted": inserted,
            "errors": errors,
        }
        with session_scope_writer() as s:
            s.execute(_UPDATE_METADATA_SQL, {
                "id": record.id,
                "metadata": json.dumps(new_metadata, ensure_ascii=False),
                "row_count": inserted,
                "parser_status": "ok",
                "parser_error": None,
            })

        log.info(
            f"ald_rayvac {record.file_name}: +{inserted} rows ({errors} errors)"
        )
        return {
            "status": "ok",
            "inserted": inserted,
            "errors": errors,
        }

    except Exception as e:
        error_msg = str(e)
        log.error(
            f"ald_rayvac {record.file_name} parse failed: {error_msg}",
            exc_info=True,
        )
        with session_scope_writer() as s:
            s.execute(_UPDATE_METADATA_SQL, {
                "id": record.id,
                "metadata": json.dumps(record.metadata or {}, ensure_ascii=False),
                "row_count": inserted,
                "parser_status": "error",
                "parser_error": error_msg[:1000],
            })
        return {
            "status": "error",
            "error": error_msg,
            "inserted": inserted,
            "errors": errors,
        }
