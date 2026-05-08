"""ALD NCD xlsx 파서 — 두 시트 → ald_ncd_runs.

NCD xlsx 구조:
- '레시피 및 결과(TTIP)' 시트: R1=그룹헤더, R2=컬럼명, R3부터 데이터, 43 컬럼
- '레시피 및 결과(TDMAT)' 시트: 같은 구조
- 첫 컬럼('공정 번호')이 비면 → 데이터 끝
- 같은 batch_no가 여러 row일 수 있음 (재공정 케이스, xlsx 그대로 보존)

처리 정책:
- 새 sha (metadata 비어있음) → 두 시트 모두 처리
- 같은 sha (metadata에 'all_processed') → skip
- xlsx 변경 시 새 source_file_id로 새 row INSERT, 옛 데이터 그대로 보존
- 검증 실패 row → parse_errors 격리 (GPT 친화 자연어 메시지)

멱등성:
- ald_ncd_runs UNIQUE (source_file_id, row_number) → ON CONFLICT DO NOTHING
- parse_errors UNIQUE (source_file_id, row_number, error_type) → ON CONFLICT DO NOTHING
"""

import json
import logging
from datetime import datetime, date
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from sqlalchemy import text

from shared.db import session_scope_writer
from etl_worker.jobs.scan_files import SourceFileRecord

log = logging.getLogger("etl.parsers.ald_ncd")

# NCD 시트 → chemistry 매핑
NCD_SHEETS = [
    ("레시피 및 결과(TTIP)", "TTIP"),
    ("레시피 및 결과(TDMAT)", "TDMAT"),
]

# NCD R2 헤더 (0-indexed 컬럼 위치 → init.sql 컬럼명)
# C20 'HT04 Temp', C22 '전체 Max-min', C24 '층 내 표준 편차',
# C25~C29 '1F~5F_Data-Average(A)', C30~C34 '1F~5F_Max-min(%)' 등은 raw_json만 보존
NCD_COL_MAP = {
    0:  'batch_no_raw',           # 공정 번호
    1:  'process_date_raw',       # Date
    2:  'temp_c',                 # Temp (°C)
    3:  'pre_heat_delay_s',       # Pre-heat delay (sec)
    4:  'stable_time_s',          # Stable time (sec)
    5:  'pre_heat_temp_c',        # Pre-heat Temp (°C)
    6:  'precursor_temp_c',       # TTIP/TDMAT Temp (°C)
    7:  'precursor_pulse_s',      # TTIP/TDMAT Pulse time (sec)
    8:  'precursor_purge_s',      # TTIP/TDMAT Purge time (sec)
    9:  'h2o_temp_c',             # H2O Temp (°C)
    10: 'h2o_pulse_s',            # H2O Pulse time (sec)
    11: 'h2o_purge_s',            # H2O Purge time (sec)
    12: 'precursor_assist_flow_sccm',  # TTIP/TDMAT Assist Flow (sccm)
    13: 'source_carrier_flow_sccm',    # Source Carrier Flow (sccm)
    14: 'h2o_carrier_flow_sccm',       # H2O Carrier Flow (sccm)
    15: 'outer_flow_sccm',             # Outer Flow (sccm)
    16: 'cycles',                      # Cycle (#)
    17: 'precursor_cum_cycles',        # TTIP/TDMAT 소비 사이클
    18: 'h2o_cum_cycles',              # H2O 소비 사이클
    19: 'chamber_clean_cum_cycles',    # 챔버 클리닝 이후 누적 사이클
    21: 'gpc_a_per_cycle',             # GPC (A/cycle)
    23: 'avg_max_min_pct',             # AVG Max-min(%)
}


def _to_float(v):
    """REAL 컬럼용 안전 변환. 변환 실패 시 None 반환 (raw 보존은 raw_json)."""
    if v is None or v == '':
        return None
    try:
        f = float(v)
        # 한글 문자열 등 비숫자가 isinstance check 우회 가능, NaN/Inf 차단
        if f != f or f == float('inf') or f == float('-inf'):
            return None
        return f
    except (ValueError, TypeError):
        return None


def _to_int(v):
    """INTEGER 컬럼용 안전 변환."""
    if v is None or v == '':
        return None
    try:
        # float 거쳐서 int (예: '150.0' → 150)
        return int(float(v))
    except (ValueError, TypeError):
        return None


def _to_date(v):
    """DATE 컬럼용 안전 변환. datetime → date, 실패 시 None."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _serialize(v):
    """JSON 직렬화 헬퍼 — datetime/date → ISO string."""
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return v


def _row_to_full_dict(row: tuple, header: list) -> dict:
    """xlsx row tuple → 전체 dict (raw_json 보존용).

    header가 None인 컬럼은 'col_N' 형태로 키 부여.
    """
    result = {}
    for i, val in enumerate(row):
        key = header[i] if i < len(header) and header[i] else f'col_{i}'
        # 키에 줄바꿈/공백 정리
        if isinstance(key, str):
            key = key.replace('\n', ' ').strip()
        result[key] = val
    return result


_INSERT_NCD_SQL = text("""
    INSERT INTO vo2.ald_ncd_runs (
        batch_no, chemistry, process_date,
        temp_c, pre_heat_delay_s, stable_time_s, pre_heat_temp_c,
        precursor_temp_c, precursor_pulse_s, precursor_purge_s,
        h2o_temp_c, h2o_pulse_s, h2o_purge_s,
        precursor_assist_flow_sccm, source_carrier_flow_sccm,
        h2o_carrier_flow_sccm, outer_flow_sccm,
        cycles, precursor_cum_cycles, h2o_cum_cycles, chamber_clean_cum_cycles,
        gpc_a_per_cycle, avg_max_min_pct,
        source_file_id, row_number, raw_json, parse_status
    ) VALUES (
        :batch_no, :chemistry, :process_date,
        :temp_c, :pre_heat_delay_s, :stable_time_s, :pre_heat_temp_c,
        :precursor_temp_c, :precursor_pulse_s, :precursor_purge_s,
        :h2o_temp_c, :h2o_pulse_s, :h2o_purge_s,
        :precursor_assist_flow_sccm, :source_carrier_flow_sccm,
        :h2o_carrier_flow_sccm, :outer_flow_sccm,
        :cycles, :precursor_cum_cycles, :h2o_cum_cycles, :chamber_clean_cum_cycles,
        :gpc_a_per_cycle, :avg_max_min_pct,
        :source_file_id, :row_number, CAST(:raw_json AS JSONB), :parse_status
    )
    ON CONFLICT (source_file_id, row_number) DO NOTHING
""")

_INSERT_PARSE_ERROR_SQL = text("""
    INSERT INTO vo2.parse_errors (
        source_file_id, row_number, error_type, error_detail, raw_data
    ) VALUES (
        :source_file_id, :row_number, :error_type, :error_detail,
        CAST(:raw_data AS JSONB)
    )
    ON CONFLICT (source_file_id, row_number, error_type) DO NOTHING
""")

_UPDATE_METADATA_SQL = text("""
    UPDATE vo2.source_files
    SET metadata = CAST(:metadata AS JSONB),
        row_count = :row_count,
        parser_status = :parser_status,
        parser_error = :parser_error,
        last_indexed_at = NOW()
    WHERE id = :id
""")


def _validate_row(row: tuple, chemistry: str, row_idx: int) -> Optional[dict]:
    """row 검증. 격리할 에러가 있으면 dict 반환, 없으면 None.

    error_detail은 GPT/운영자 친화 자연어 (어떻게 해결할지 명시).
    """
    batch_raw = row[0] if len(row) > 0 else None
    date_raw = row[1] if len(row) > 1 else None

    # 1. batch_no 정수 변환 검증
    try:
        int(batch_raw) if batch_raw is not None else None
        if batch_raw is None:
            # 빈 row는 호출 전에 break됨 — 여기 도달 안 함, 안전망
            return None
    except (ValueError, TypeError):
        return {
            "error_type": "batch_no_invalid",
            "error_detail": (
                f"NCD '{chemistry}' 시트 R{row_idx} row의 공정 번호가 "
                f"'{batch_raw}'로 정수가 아닙니다. ald_ncd_runs에 INSERT하지 못했습니다. "
                f"운영자가 xlsx의 R{row_idx} 첫 컬럼을 정확한 정수로 수정하면 "
                f"다음 5분 tick에서 자동 처리됩니다 (단 이 row는 새 source_file_id로 들어가니 "
                f"옛 source_file_id의 격리 row는 그대로 남습니다)."
            ),
        }

    # 2. Date 검증 (process_date NOT NULL)
    if not isinstance(date_raw, (datetime, date)):
        return {
            "error_type": "date_missing",
            "error_detail": (
                f"NCD '{chemistry}' 시트 R{row_idx} row의 Date 컬럼이 비어있거나 "
                f"datetime이 아닙니다 (실제 값: {date_raw!r}). "
                f"ald_ncd_runs.process_date는 NOT NULL이라 INSERT하지 못했습니다. "
                f"운영자가 xlsx에 Date를 입력하면 다음 tick에서 자동 처리됩니다."
            ),
        }

    return None


def _build_payload(
    row: tuple, chemistry: str, source_file_id: int, row_number: int, header: list
) -> dict:
    """검증 통과한 row → ald_ncd_runs INSERT payload."""
    batch_no = _to_int(row[0])
    process_date = _to_date(row[1])

    # 매핑된 컬럼 추출
    payload = {
        'batch_no': batch_no,
        'chemistry': chemistry,
        'process_date': process_date,
        'temp_c': _to_float(row[2]) if len(row) > 2 else None,
        'pre_heat_delay_s': _to_float(row[3]) if len(row) > 3 else None,
        'stable_time_s': _to_float(row[4]) if len(row) > 4 else None,
        'pre_heat_temp_c': _to_float(row[5]) if len(row) > 5 else None,
        'precursor_temp_c': _to_float(row[6]) if len(row) > 6 else None,
        'precursor_pulse_s': _to_float(row[7]) if len(row) > 7 else None,
        'precursor_purge_s': _to_float(row[8]) if len(row) > 8 else None,
        'h2o_temp_c': _to_float(row[9]) if len(row) > 9 else None,
        'h2o_pulse_s': _to_float(row[10]) if len(row) > 10 else None,
        'h2o_purge_s': _to_float(row[11]) if len(row) > 11 else None,
        'precursor_assist_flow_sccm': _to_float(row[12]) if len(row) > 12 else None,
        'source_carrier_flow_sccm': _to_float(row[13]) if len(row) > 13 else None,
        'h2o_carrier_flow_sccm': _to_float(row[14]) if len(row) > 14 else None,
        'outer_flow_sccm': _to_float(row[15]) if len(row) > 15 else None,
        'cycles': _to_int(row[16]) if len(row) > 16 else None,
        'precursor_cum_cycles': _to_int(row[17]) if len(row) > 17 else None,
        'h2o_cum_cycles': _to_int(row[18]) if len(row) > 18 else None,
        'chamber_clean_cum_cycles': _to_int(row[19]) if len(row) > 19 else None,
        'gpc_a_per_cycle': _to_float(row[21]) if len(row) > 21 else None,
        'avg_max_min_pct': _to_float(row[23]) if len(row) > 23 else None,
        'source_file_id': source_file_id,
        'row_number': row_number,
        'parse_status': None,
    }

    # raw_json: 모든 원본 컬럼 보존 (분석 안전망)
    raw = _row_to_full_dict(row, header)
    payload['raw_json'] = json.dumps(
        {k: _serialize(v) for k, v in raw.items()},
        ensure_ascii=False,
    )

    return payload


def _process_sheet(
    ws, sheet_name: str, chemistry: str, source_file_id: int
) -> tuple[int, int]:
    """한 시트 처리. (inserted_count, error_count) 반환."""
    inserted = 0
    errors = 0

    # R2 헤더 추출 (raw_json 키로 사용)
    header_rows = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    header = list(header_rows[0]) if header_rows else []
    # 헤더 정리 (None 제거, 줄바꿈 정리)
    header = [
        (h.replace('\n', ' ').strip() if isinstance(h, str) else h)
        for h in header
    ]

    with session_scope_writer() as s:
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=3, values_only=True), start=3
        ):
            first_cell = row[0] if row else None

            # 빈 row 만나면 break (xlsx 데이터 끝)
            if first_cell is None or (
                isinstance(first_cell, str) and not first_cell.strip()
            ):
                break

            # 검증
            error = _validate_row(row, chemistry, row_idx)
            if error:
                # parse_errors 격리
                raw_data = _row_to_full_dict(row, header)
                s.execute(_INSERT_PARSE_ERROR_SQL, {
                    "source_file_id": source_file_id,
                    "row_number": row_idx,
                    "error_type": error["error_type"],
                    "error_detail": error["error_detail"],
                    "raw_data": json.dumps(
                        {k: _serialize(v) for k, v in raw_data.items()},
                        ensure_ascii=False,
                    ),
                })
                errors += 1
                continue

            # 정상 row INSERT
            payload = _build_payload(row, chemistry, source_file_id, row_idx, header)
            s.execute(_INSERT_NCD_SQL, payload)
            inserted += 1

    log.info(
        f"NCD '{chemistry}' 시트 처리: +{inserted} rows, {errors} errors"
    )
    return inserted, errors


def parse_ald_ncd(record: SourceFileRecord) -> dict:
    """ALD NCD xlsx 메인 진입.

    Returns:
        {"status": "ok"|"error"|"skipped",
         "ttip_inserted": N, "tdmat_inserted": M,
         "ttip_errors": X, "tdmat_errors": Y, ...}
    """
    if record.is_race_unsafe:
        log.info(
            f"skip {record.file_name} (race_unsafe, mtime too recent)"
        )
        return {
            "status": "skipped", "reason": "race_unsafe",
            "ttip_inserted": 0, "tdmat_inserted": 0,
            "ttip_errors": 0, "tdmat_errors": 0,
        }

    # 같은 sha 이미 처리된 경우 skip (성능 절약)
    if record.metadata and record.metadata.get("all_processed"):
        log.info(
            f"skip {record.file_name} (sha already processed)"
        )
        return {
            "status": "skipped", "reason": "already_processed",
            "ttip_inserted": 0, "tdmat_inserted": 0,
            "ttip_errors": 0, "tdmat_errors": 0,
        }

    ttip_inserted = 0
    tdmat_inserted = 0
    ttip_errors = 0
    tdmat_errors = 0

    try:
        wb = load_workbook(record.file_path, read_only=True, data_only=True)
        try:
            for sheet_name, chemistry in NCD_SHEETS:
                if sheet_name not in wb.sheetnames:
                    log.warning(
                        f"sheet '{sheet_name}' not found in {record.file_name}"
                    )
                    continue
                ws = wb[sheet_name]
                inserted, errors = _process_sheet(
                    ws, sheet_name, chemistry, record.id
                )
                if chemistry == "TTIP":
                    ttip_inserted, ttip_errors = inserted, errors
                else:
                    tdmat_inserted, tdmat_errors = inserted, errors
        finally:
            wb.close()

        # metadata 갱신: 처리 완료 표시
        new_metadata = {
            "all_processed": True,
            "ttip_inserted": ttip_inserted,
            "tdmat_inserted": tdmat_inserted,
            "ttip_errors": ttip_errors,
            "tdmat_errors": tdmat_errors,
        }
        with session_scope_writer() as s:
            s.execute(_UPDATE_METADATA_SQL, {
                "id": record.id,
                "metadata": json.dumps(new_metadata, ensure_ascii=False),
                "row_count": ttip_inserted + tdmat_inserted,
                "parser_status": "ok",
                "parser_error": None,
            })

        log.info(
            f"ald_ncd {record.file_name}: "
            f"TTIP +{ttip_inserted} ({ttip_errors} err), "
            f"TDMAT +{tdmat_inserted} ({tdmat_errors} err)"
        )
        return {
            "status": "ok",
            "ttip_inserted": ttip_inserted,
            "tdmat_inserted": tdmat_inserted,
            "ttip_errors": ttip_errors,
            "tdmat_errors": tdmat_errors,
        }

    except Exception as e:
        error_msg = str(e)
        log.error(
            f"ald_ncd {record.file_name} parse failed: {error_msg}",
            exc_info=True,
        )
        with session_scope_writer() as s:
            s.execute(_UPDATE_METADATA_SQL, {
                "id": record.id,
                "metadata": json.dumps(record.metadata or {}, ensure_ascii=False),
                "row_count": ttip_inserted + tdmat_inserted,
                "parser_status": "error",
                "parser_error": error_msg[:1000],
            })
        return {
            "status": "error",
            "error": error_msg,
            "ttip_inserted": ttip_inserted,
            "tdmat_inserted": tdmat_inserted,
            "ttip_errors": ttip_errors,
            "tdmat_errors": tdmat_errors,
        }
